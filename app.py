
import re
import pandas as pd
import streamlit as st
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Model Number / SKU → Items Matcher (Preserve Format)", layout="wide")
st.title("🔗 Model Number / SKU → Items Matcher · Preserve Original Formatting (v3.7)")

with st.expander("📘 Usage Instructions (click to expand)"):
    st.markdown("""
**What this tool does**
- Edits your uploaded **File A** *in place* to preserve formatting (styles, widths, filters, freeze panes, merged cells).
- Adds a new **Items** column **at the end of each sheet** that has a header exactly **`Model Number`** or **`SKU`** (case-insensitive).  
  *Note:* Columns like **`FN SKU`**, **`FNSKU`**, etc. are **not** matched.

**How to use**
1. **Upload File A** — Excel to be matched (multi-sheet supported).
2. **Upload File B** — Mapping file with **2 columns**: `Model Number`, `Items`.
3. Click **Run Matching (Preserve Format)**.
4. Download the result named **`<FileA>_update.xlsx`**.

**Matching rules**
- Split multiple models in a cell by comma (`,` or `，`), and process each token.
- Ignore leading quantities like `49 x ...`, `289*...`.
- For each token, find the **longest** `Model Number` in File B that is **contained** in the token (case-insensitive).
- If a token has no match → `N/A`. If any token is `N/A`, the **entire row** is highlighted red.

**Formatting rules**
- The new **Items** column is **appended to the last column** to avoid shifting your existing columns/merge/formulas.
- Header and body styles of the new column are copied from the detected header column so borders align.
- Filter ranges are widened **horizontally only** to include the new column. Bottom blank rows are not touched.
""")

with st.sidebar:
    st.header("Upload files (Drag & Drop supported)")
    file_a = st.file_uploader("Upload A (Excel to be matched)", type=["xlsx"])
    file_b = st.file_uploader("Upload B (Mapping: 2 columns)", type=["xlsx", "xls"])
    run = st.button("🚀 Run Matching (Preserve Format)", use_container_width=True)

def _clean_leading_qty(txt: str) -> str:
    return re.sub(r"^\s*\d+\s*[x\*]\s*", "", txt, flags=re.IGNORECASE)

def _split_models(cell: str):
    parts = re.split(r"[,\，]", cell)
    return [p.strip() for p in parts if p.strip()]

def _norm_cols(cols):
    return [str(c).strip() for c in cols]

def build_mapping(b_df: pd.DataFrame):
    b_df = b_df.copy()
    b_df.columns = _norm_cols(b_df.columns)
    col1 = None; col2 = None
    for c in b_df.columns:
        if c.lower() == "model number" and col1 is None:
            col1 = c
        if c.lower() == "items" and col2 is None:
            col2 = c
    if col1 is None or col2 is None:
        col1, col2 = b_df.columns[:2]
    tmp = b_df[[col1, col2]].copy()
    tmp[col1] = tmp[col1].astype(str).str.strip()
    tmp[col2] = tmp[col2].astype(str).str.strip()
    tmp = tmp[tmp[col1] != ""]
    mapping = dict(zip(tmp[col1], tmp[col2]))
    keys_sorted = sorted(mapping.keys(), key=lambda x: len(str(x)), reverse=True)
    keys_sorted_lower = [str(k).lower() for k in keys_sorted]
    return mapping, keys_sorted, keys_sorted_lower

def longest_substring_match(token: str, keys_sorted, keys_sorted_lower):
    t = token.lower()
    for k_lower, k in zip(keys_sorted_lower, keys_sorted):
        if k_lower in t:
            return k
    return None

def last_data_row_in_column(ws, col_idx: int, start_row: int = 2) -> int:
    for r in range(ws.max_row, start_row - 1, -1):
        val = ws.cell(row=r, column=col_idx).value
        if val not in (None, ""):
            return r
    return start_row - 1

def process_workbook(a_bytes: bytes, b_df: pd.DataFrame) -> bytes:
    mapping, keys_sorted, keys_sorted_lower = build_mapping(b_df)
    red_fill = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid")

    wb = load_workbook(BytesIO(a_bytes))
    for ws in wb.worksheets:
        header_row = 1
        key_col_idx = None
        header_style = None

        # ✅ EXACT header match: "model number" OR "sku" only
        for cell in ws[header_row]:
            if not cell.value:
                continue
            v = str(cell.value).strip().lower()
            if v == "model number" or v == "sku":
                key_col_idx = cell.column
                header_style = cell._style
                break

        if key_col_idx is None:
            continue

        # Copy one body style from the second row of the key column
        body_style = None
        if ws.max_row >= header_row + 1:
            body_style = ws.cell(row=header_row + 1, column=key_col_idx)._style

        # Append Items as last column
        insert_at = ws.max_column + 1
        header_cell = ws.cell(row=header_row, column=insert_at)
        header_cell.value = "Items"
        try:
            header_cell._style = header_style
        except Exception:
            pass
        ws.column_dimensions[get_column_letter(insert_at)].width = (
            ws.column_dimensions[get_column_letter(key_col_idx)].width or 15
        )

        last_row = last_data_row_in_column(ws, key_col_idx, start_row=header_row + 1)
        for r in range(header_row + 1, last_row + 1):
            raw = ws.cell(row=r, column=key_col_idx).value
            text = (str(raw).strip()) if raw is not None else ""
            if text == "":
                continue
            cleaned = _clean_leading_qty(text)
            parts = _split_models(cleaned)
            out_items = []
            has_na = False
            for p in parts:
                mk = longest_substring_match(p, keys_sorted, keys_sorted_lower)
                if mk is None:
                    out_items.append("N/A"); has_na = True
                else:
                    val = mapping.get(mk, "N/A")
                    out_items.append(val)
                    if val == "N/A":
                        has_na = True
            tgt = ws.cell(row=r, column=insert_at)
            tgt.value = ",".join(out_items)
            try:
                if body_style is not None:
                    tgt._style = body_style
            except Exception:
                pass
            if has_na:
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c).fill = red_fill

        # Widen filter range horizontally only (if a filter exists)
        try:
            if ws.auto_filter and ws.auto_filter.ref:
                ref = ws.auto_filter.ref
                # keep original rows; only extend the end column
                import re as _re
                m1 = _re.match(r"([A-Z]{1,3})(\d+):([A-Z]{1,3})(\d+)", ref or "")
                if m1:
                    sc, sr, ec, er = m1.groups()
                    ws.auto_filter.ref = f"{sc}{sr}:{get_column_letter(ws.max_column)}{er}"
        except Exception:
            pass

    out = BytesIO()
    wb.save(out)
    return out.getvalue()

if run:
    if not file_a or not file_b:
        st.error("Please upload both A and B files.")
        st.stop()
    try:
        b_df = pd.read_excel(file_b)
        processed = process_workbook(file_a.read(), b_df)
        base_name = file_a.name.rsplit(".", 1)[0] if file_a.name else "A_matched"
        out_name = f"{base_name}_update.xlsx"
        st.success("Done! Items column appended at the end. Exact header match: 'Model Number' or 'SKU'.")
        st.download_button("⬇️ Download result", data=processed,
                           file_name=out_name,
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        st.exception(e)
else:
    st.info("Drag & drop your A and B files on the left, then click the button.")
